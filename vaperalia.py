from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from dotenv import load_dotenv
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font
import os

load_dotenv()
base_dir = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(base_dir, 'static')

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Nombre'
ws['B1'] = 'Imagen'
ws['C1'] = 'Link'
ws['D1'] = 'SKU'
ws['E1'] = 'Referencia'
font_azul = Font(color='0000FF') 

def iniciar_chrome():
    ruta = f"{base_dir}\chromedriver.exe"
    options = Options()
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0"
    options.add_argument(f"user-agent={user_agent}")
    options.add_argument("--start-maximized")
    #options.add_argument("--window-size=970,1000")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-extensions")    
    options.add_argument("--disable-notifications")    
    options.add_argument("--ignore-certificate-errors")    
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--no-proxy-server")
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    exp_opt = [
        "enable-automaiton",
        "ignore-certificate-errors",
        "enable-logging"
    ]
    options.add_experimental_option("excludeSwitches", exp_opt)
    
    prefs = {
        "profile.default_content_setting_values.notifications": 2,
        "intl.accept_langiages": ["es-ES", "es"],
        "credentials_enable_service": False
    }
    options.add_experimental_option("prefs", prefs)
    
    s = Service(ruta)
    driver = webdriver.Chrome(service=s, options=options)
    #driver.set_window_position(0,0)
    return driver


def scrape_page():
    url = "https://vaperalia.es/nuevos-productos"
    driver = iniciar_chrome()
    driver.get(url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//h2[contains(text(), 'Novedades')]"))
    )
    productos = driver.find_elements(By.CSS_SELECTOR, "div.product-container")
    print(len(productos))
    counter = 2
    for p in productos:
        imagen = driver.find_element(By.CSS_SELECTOR, "img.replace-2x.img-responsive").get_attribute("src")
        nombre_obj = driver.find_element(By.CSS_SELECTOR, "a.product-name")
        nombre = nombre_obj.text
        link = nombre_obj.get_attribute("href")
        driver.get(link)
        
        sku = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span[itemprop='ean13']"))
        ).text
        referencia = driver.find_element(By.CSS_SELECTOR, "span[itemprop='sku']").text
        print(f"Producto con referencia {referencia} scrapeado.")
        
        ws.cell(row=counter, column=1).value = nombre
        
        ws.cell(row=counter, column=2).value = imagen
        ws.cell(row=counter, column=2).hyperlink = imagen
        ws.cell(row=counter, column=2).font = font_azul
        
        ws.cell(row=counter, column=3).value = link
        ws.cell(row=counter, column=3).hyperlink = link
        ws.cell(row=counter, column=3).font = font_azul
        
        ws.cell(row=counter, column=4).value = sku
        ws.cell(row=counter, column=5).value = referencia
        
        counter+=1

    wb.save(os.path.join(static_dir, 'productos_vaperalia.xlsx'))
    print("Datos guardados en productos_vaperalia.xlsx")
    driver.quit()


if __name__ == "__main__":
    scrape_page()
    