import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import requests, os

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Nombre'
ws['B1'] = 'Descripción'
ws['C1'] = 'Imagen'
ws['D1'] = 'Link'
ws['E1'] = 'SKU'
font_azul = Font(color='0000FF') 

url = "https://nueva.eciglogistica.com/novedades"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
}

def scrape_page():
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        productos = soup.find_all("div", class_="product card-product box")
        print(len(productos))
        productos_info = []
        base_dir = os.path.dirname(os.path.abspath(__file__))
        static_dir = os.path.join(base_dir, 'static')
        if not os.path.exists(static_dir):
            os.makedirs(static_dir)

        counter = 2
        for producto in productos:
            nombre = producto.find("h5").text.strip()
            descripcion = producto.find("div", class_="product-body").text.strip()
            imagen = producto.find("img", class_="img-fluid")["src"]
            link = producto.find("a", class_="product-header")["href"]
            sku = ""
            response = requests.get(link, headers=headers)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                sku = soup.find("p", class_="grey-texts mt-lg-0 mt-2 mb-2").text.strip().split(":")[1]
                print(f"Artículo {sku} scrapeado.")
            else:
                print(f"Error al entrar al link: {link}")
                return False
            
            ws.cell(row=counter, column=1).value = nombre
            ws.cell(row=counter, column=2).value = descripcion
            ws.cell(row=counter, column=3).value = imagen
            ws.cell(row=counter, column=3).hyperlink = imagen
            ws.cell(row=counter, column=3).font = font_azul
            ws.cell(row=counter, column=4).value = link
            ws.cell(row=counter, column=4).hyperlink = link
            ws.cell(row=counter, column=4).font = font_azul
            ws.cell(row=counter, column=5).value = sku
            
            p_info = {"nombre": nombre, "descripcion": descripcion, "imagen": imagen, "link": link, "sku": sku}
            productos_info.append(p_info)

            counter+=1

        wb.save(os.path.join(static_dir, 'productos.xlsx'))
        print("Datos guardados en productos.xlsx")
        return productos_info
    else:
        print("Error al realizar la solicitud")
        return False
